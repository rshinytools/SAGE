"""
Audit Service Module
====================

High-level service for audit logging with export capabilities.
Provides a simple API for logging events throughout the application.
"""

import os
import json
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple

from .models import (
    AuditAction,
    AuditStatus,
    AuditEvent,
    QueryAuditDetails,
    AuditLog,
    AuditFilters,
    AuditStatistics,
    ElectronicSignature,
    IntegrityCheckResult,
)
from .database import AuditDB


class AuditService:
    """
    Centralized audit logging service with 21 CFR Part 11 compliance.

    Usage:
        service = get_audit_service()
        audit_id = service.log_login("user123", "Dr. Smith", "192.168.1.1")
        service.log_query(audit_id, query_details)
    """

    # Maximum size for response body logging (bytes)
    MAX_RESPONSE_SIZE = int(os.getenv("AUDIT_MAX_RESPONSE_SIZE", "10000"))

    # Paths to exclude from API request logging
    EXCLUDED_PATHS = os.getenv(
        "AUDIT_EXCLUDED_PATHS",
        "/health,/docs,/openapi.json,/redoc,/favicon.ico"
    ).split(",")

    def __init__(self, db_path: Optional[str] = None):
        """Initialize the audit service."""
        self._db = AuditDB(db_path)

    # ==================== AUTHENTICATION LOGGING ====================

    def log_login(
        self,
        user_id: str,
        username: str,
        ip_address: Optional[str] = None,
        user_agent: Optional[str] = None,
        success: bool = True,
        failure_reason: Optional[str] = None,
    ) -> int:
        """
        Log a login attempt.

        Args:
            user_id: The user ID.
            username: The username or display name.
            ip_address: Client IP address.
            user_agent: Client user agent string.
            success: Whether login was successful.
            failure_reason: Reason for failure if not successful.

        Returns:
            The audit log ID.
        """
        event = AuditEvent(
            user_id=user_id,
            username=username,
            action=AuditAction.LOGIN if success else AuditAction.LOGIN_FAILED,
            status=AuditStatus.SUCCESS if success else AuditStatus.FAILURE,
            ip_address=ip_address,
            user_agent=user_agent,
            error_message=failure_reason,
            details={"method": "jwt"} if success else {"failure_reason": failure_reason},
        )
        return self._db.insert_log(event)

    def log_logout(
        self,
        user_id: str,
        username: str,
        ip_address: Optional[str] = None,
        session_duration_seconds: Optional[int] = None,
    ) -> int:
        """Log a logout event."""
        event = AuditEvent(
            user_id=user_id,
            username=username,
            action=AuditAction.LOGOUT,
            status=AuditStatus.SUCCESS,
            ip_address=ip_address,
            details={"session_duration_seconds": session_duration_seconds} if session_duration_seconds else None,
        )
        return self._db.insert_log(event)

    def log_token_refresh(
        self,
        user_id: str,
        username: str,
        ip_address: Optional[str] = None,
    ) -> int:
        """Log a token refresh event."""
        event = AuditEvent(
            user_id=user_id,
            username=username,
            action=AuditAction.TOKEN_REFRESH,
            status=AuditStatus.SUCCESS,
            ip_address=ip_address,
        )
        return self._db.insert_log(event)

    # ==================== QUERY LOGGING ====================

    def log_query(
        self,
        user_id: str,
        username: str,
        question: str,
        success: bool = True,
        error_message: Optional[str] = None,
        ip_address: Optional[str] = None,
        duration_ms: Optional[int] = None,
        resource_id: Optional[str] = None,
    ) -> int:
        """
        Log a query event.

        Args:
            user_id: The user ID.
            username: The username.
            question: The original question.
            success: Whether query was successful.
            error_message: Error message if failed.
            ip_address: Client IP address.
            duration_ms: Query duration in milliseconds.
            resource_id: Conversation or session ID.

        Returns:
            The audit log ID.
        """
        event = AuditEvent(
            user_id=user_id,
            username=username,
            action=AuditAction.QUERY if success else AuditAction.QUERY_FAILED,
            resource_type="chat",
            resource_id=resource_id,
            status=AuditStatus.SUCCESS if success else AuditStatus.FAILURE,
            ip_address=ip_address,
            duration_ms=duration_ms,
            error_message=error_message,
            details={"question_preview": question[:200]} if question else None,
        )
        return self._db.insert_log(event)

    def log_query_details(self, audit_log_id: int, details: QueryAuditDetails) -> int:
        """
        Add detailed query information to an existing audit log.

        Args:
            audit_log_id: The ID of the parent audit log.
            details: The query details.

        Returns:
            The query details record ID.
        """
        return self._db.insert_query_details(audit_log_id, details)

    # ==================== DATA OPERATION LOGGING ====================

    def log_data_upload(
        self,
        user_id: str,
        username: str,
        filename: str,
        file_size: int,
        row_count: Optional[int] = None,
        columns: Optional[List[str]] = None,
        ip_address: Optional[str] = None,
        duration_ms: Optional[int] = None,
        success: bool = True,
        error_message: Optional[str] = None,
    ) -> int:
        """Log a data upload event."""
        event = AuditEvent(
            user_id=user_id,
            username=username,
            action=AuditAction.DATA_UPLOAD,
            resource_type="data_file",
            resource_id=filename,
            status=AuditStatus.SUCCESS if success else AuditStatus.FAILURE,
            ip_address=ip_address,
            duration_ms=duration_ms,
            error_message=error_message,
            details={
                "file_size": file_size,
                "row_count": row_count,
                "columns": columns[:20] if columns else None,  # Limit columns logged
                "column_count": len(columns) if columns else None,
            },
        )
        return self._db.insert_log(event)

    def log_data_export(
        self,
        user_id: str,
        username: str,
        export_format: str,
        row_count: int,
        ip_address: Optional[str] = None,
        resource_id: Optional[str] = None,
    ) -> int:
        """Log a data export event."""
        event = AuditEvent(
            user_id=user_id,
            username=username,
            action=AuditAction.DATA_EXPORT,
            resource_type="export",
            resource_id=resource_id,
            status=AuditStatus.SUCCESS,
            ip_address=ip_address,
            details={
                "format": export_format,
                "row_count": row_count,
            },
        )
        return self._db.insert_log(event)

    # ==================== API REQUEST LOGGING ====================

    def log_api_request(
        self,
        user_id: str,
        username: str,
        method: str,
        path: str,
        status_code: int,
        duration_ms: int,
        ip_address: Optional[str] = None,
        user_agent: Optional[str] = None,
        request_body: Optional[str] = None,
        response_body: Optional[str] = None,
        error_message: Optional[str] = None,
    ) -> Optional[int]:
        """
        Log an API request.

        Args:
            user_id: The user ID.
            username: The username.
            method: HTTP method.
            path: Request path.
            status_code: Response status code.
            duration_ms: Request duration in milliseconds.
            ip_address: Client IP address.
            user_agent: Client user agent.
            request_body: Request body (sanitized).
            response_body: Response body (truncated).
            error_message: Error message if failed.

        Returns:
            The audit log ID, or None if path is excluded.
        """
        # Skip excluded paths
        if any(path.startswith(excluded) for excluded in self.EXCLUDED_PATHS):
            return None

        # Truncate response body if too large
        if response_body and len(response_body) > self.MAX_RESPONSE_SIZE:
            response_body = response_body[:self.MAX_RESPONSE_SIZE] + "... [truncated]"

        # Determine status
        if status_code >= 500:
            status = AuditStatus.ERROR
        elif status_code >= 400:
            status = AuditStatus.FAILURE
        else:
            status = AuditStatus.SUCCESS

        event = AuditEvent(
            user_id=user_id,
            username=username,
            action=AuditAction.API_REQUEST,
            resource_type="api",
            status=status,
            ip_address=ip_address,
            user_agent=user_agent,
            request_method=method,
            request_path=path,
            request_body=request_body,
            response_status=status_code,
            response_body=response_body,
            duration_ms=duration_ms,
            error_message=error_message,
        )
        return self._db.insert_log(event)

    # ==================== SYSTEM LOGGING ====================

    def log_system_startup(self, version: str = "1.0.0", config: Optional[Dict] = None) -> int:
        """Log system startup."""
        event = AuditEvent(
            user_id="system",
            username="SAGE System",
            action=AuditAction.SYSTEM_STARTUP,
            resource_type="system",
            status=AuditStatus.SUCCESS,
            details={
                "version": version,
                "config": config,
            },
        )
        return self._db.insert_log(event)

    def log_system_shutdown(self, reason: str = "normal", uptime_seconds: Optional[int] = None) -> int:
        """Log system shutdown."""
        event = AuditEvent(
            user_id="system",
            username="SAGE System",
            action=AuditAction.SYSTEM_SHUTDOWN,
            resource_type="system",
            status=AuditStatus.SUCCESS,
            details={
                "reason": reason,
                "uptime_seconds": uptime_seconds,
            },
        )
        return self._db.insert_log(event)

    def log_config_change(
        self,
        user_id: str,
        username: str,
        config_key: str,
        old_value: Any,
        new_value: Any,
        ip_address: Optional[str] = None,
    ) -> int:
        """Log a configuration change."""
        event = AuditEvent(
            user_id=user_id,
            username=username,
            action=AuditAction.CONFIG_CHANGE,
            resource_type="config",
            resource_id=config_key,
            status=AuditStatus.SUCCESS,
            ip_address=ip_address,
            details={
                "key": config_key,
                "old_value": str(old_value)[:500],
                "new_value": str(new_value)[:500],
            },
        )
        return self._db.insert_log(event)

    # ==================== METADATA LOGGING ====================

    def log_metadata_action(
        self,
        user_id: str,
        username: str,
        action: str,  # 'approved', 'rejected', 'modified'
        variable_name: str,
        table_name: str,
        ip_address: Optional[str] = None,
        details: Optional[Dict] = None,
    ) -> int:
        """Log a metadata action."""
        action_map = {
            'approved': AuditAction.METADATA_APPROVED,
            'rejected': AuditAction.METADATA_REJECTED,
            'modified': AuditAction.METADATA_MODIFIED,
        }
        event = AuditEvent(
            user_id=user_id,
            username=username,
            action=action_map.get(action, AuditAction.METADATA_MODIFIED),
            resource_type="metadata",
            resource_id=f"{table_name}.{variable_name}",
            status=AuditStatus.SUCCESS,
            ip_address=ip_address,
            details=details,
        )
        return self._db.insert_log(event)

    # ==================== GENERIC LOGGING ====================

    def log_event(self, event: AuditEvent) -> int:
        """Log a generic audit event."""
        return self._db.insert_log(event)

    # ==================== QUERY OPERATIONS ====================

    def get_log(self, log_id: int, include_details: bool = True) -> Optional[AuditLog]:
        """Get a single audit log by ID."""
        return self._db.get_log(log_id, include_details)

    def search_logs(self, filters: AuditFilters) -> Tuple[List[AuditLog], int]:
        """Search audit logs with filters."""
        return self._db.search_logs(filters)

    def get_query_details(self, audit_log_id: int) -> Optional[QueryAuditDetails]:
        """Get query details for an audit log."""
        return self._db.get_query_details(audit_log_id)

    def get_statistics(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ) -> AuditStatistics:
        """Get audit statistics."""
        return self._db.get_statistics(start_date, end_date)

    def get_available_actions(self) -> List[str]:
        """Get list of available action types."""
        return self._db.get_available_actions()

    def get_available_users(self) -> List[Dict[str, str]]:
        """Get list of users with audit entries."""
        return self._db.get_available_users()

    def get_available_resource_types(self) -> List[str]:
        """Get list of resource types."""
        return self._db.get_available_resource_types()

    # ==================== INTEGRITY & SIGNATURES ====================

    def verify_integrity(self, log_id: int) -> IntegrityCheckResult:
        """Verify the integrity of an audit log."""
        return self._db.verify_integrity(log_id)

    def add_signature(
        self,
        audit_log_id: int,
        user_id: str,
        username: str,
        meaning: str,
    ) -> int:
        """Add an electronic signature to an audit log."""
        signature = ElectronicSignature(
            audit_log_id=audit_log_id,
            signer_user_id=user_id,
            signer_username=username,
            signature_meaning=meaning,
        )
        return self._db.insert_signature(signature)

    # ==================== EXPORT OPERATIONS ====================

    def export_to_excel(self, filters: AuditFilters, output_path: Optional[str] = None) -> str:
        """
        Export filtered audit logs to Excel.

        Args:
            filters: Search filters.
            output_path: Output file path. If None, creates temp file.

        Returns:
            Path to the generated Excel file.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        # Get all matching logs (override pagination for export)
        export_filters = filters.model_copy()
        export_filters.page = 1
        export_filters.page_size = 100000  # Get all

        logs, total = self.search_logs(export_filters)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Logs"

        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = [
            "ID", "Timestamp", "User", "Action", "Resource Type", "Resource ID",
            "Status", "IP Address", "Method", "Path", "Duration (ms)", "Error"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Data rows
        for row_idx, log in enumerate(logs, 2):
            ws.cell(row=row_idx, column=1, value=log.id).border = border
            ws.cell(row=row_idx, column=2, value=log.timestamp.isoformat()).border = border
            ws.cell(row=row_idx, column=3, value=log.username).border = border
            ws.cell(row=row_idx, column=4, value=log.action).border = border
            ws.cell(row=row_idx, column=5, value=log.resource_type).border = border
            ws.cell(row=row_idx, column=6, value=log.resource_id).border = border
            ws.cell(row=row_idx, column=7, value=log.status).border = border
            ws.cell(row=row_idx, column=8, value=log.ip_address).border = border
            ws.cell(row=row_idx, column=9, value=log.request_method).border = border
            ws.cell(row=row_idx, column=10, value=log.request_path).border = border
            ws.cell(row=row_idx, column=11, value=log.duration_ms).border = border
            ws.cell(row=row_idx, column=12, value=log.error_message).border = border

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Generate output path if not provided
        if output_path is None:
            date_range = ""
            if filters.start_date:
                date_range = f"_{filters.start_date.strftime('%Y%m%d')}"
            if filters.end_date:
                date_range += f"_to_{filters.end_date.strftime('%Y%m%d')}"
            output_path = tempfile.mktemp(suffix=f"_audit_logs{date_range}.xlsx")

        wb.save(output_path)
        return output_path

    def export_to_pdf(self, filters: AuditFilters, output_path: Optional[str] = None) -> str:
        """
        Export filtered audit logs to PDF.

        Args:
            filters: Search filters.
            output_path: Output file path. If None, creates temp file.

        Returns:
            Path to the generated PDF file.
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        except ImportError:
            raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")

        # Get statistics and logs
        stats = self.get_statistics(filters.start_date, filters.end_date)

        export_filters = filters.model_copy()
        export_filters.page = 1
        export_filters.page_size = 1000  # Limit for PDF
        logs, total = self.search_logs(export_filters)

        # Generate output path if not provided
        if output_path is None:
            date_range = ""
            if filters.start_date:
                date_range = f"_{filters.start_date.strftime('%Y%m%d')}"
            if filters.end_date:
                date_range += f"_to_{filters.end_date.strftime('%Y%m%d')}"
            output_path = tempfile.mktemp(suffix=f"_audit_report{date_range}.pdf")

        doc = SimpleDocTemplate(output_path, pagesize=landscape(letter))
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
        )
        elements.append(Paragraph("SAGE Audit Report", title_style))

        # Report metadata
        meta_style = styles['Normal']
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", meta_style))
        if filters.start_date or filters.end_date:
            date_range_text = f"Date Range: {filters.start_date or 'Beginning'} to {filters.end_date or 'Now'}"
            elements.append(Paragraph(date_range_text, meta_style))
        elements.append(Spacer(1, 20))

        # Statistics summary
        elements.append(Paragraph("Summary Statistics", styles['Heading2']))
        stats_data = [
            ["Total Events", str(stats.total_events)],
            ["Successful", str(stats.by_status.get('success', 0))],
            ["Failed", str(stats.by_status.get('failure', 0))],
            ["Errors", str(stats.by_status.get('error', 0))],
        ]
        if stats.average_query_confidence:
            stats_data.append(["Avg Query Confidence", f"{stats.average_query_confidence}%"])
        if stats.average_duration_ms:
            stats_data.append(["Avg Duration", f"{stats.average_duration_ms}ms"])

        stats_table = Table(stats_data, colWidths=[150, 100])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 30))

        # Audit log table
        elements.append(Paragraph(f"Audit Log Entries (showing {len(logs)} of {total})", styles['Heading2']))

        # Table headers and data
        table_data = [["Timestamp", "User", "Action", "Status", "Resource", "Duration"]]
        for log in logs:
            table_data.append([
                log.timestamp.strftime('%Y-%m-%d %H:%M'),
                log.username[:20],
                log.action[:20],
                log.status,
                (log.resource_type or '')[:15],
                f"{log.duration_ms}ms" if log.duration_ms else "",
            ])

        log_table = Table(table_data, colWidths=[100, 100, 100, 60, 100, 60])
        log_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        ]))
        elements.append(log_table)

        # Build PDF
        doc.build(elements)
        return output_path

    def export_to_csv(self, filters: AuditFilters, output_path: Optional[str] = None) -> str:
        """
        Export filtered audit logs to CSV.

        Args:
            filters: Search filters.
            output_path: Output file path. If None, creates temp file.

        Returns:
            Path to the generated CSV file.
        """
        import csv

        # Get all matching logs
        export_filters = filters.model_copy()
        export_filters.page = 1
        export_filters.page_size = 100000

        logs, total = self.search_logs(export_filters)

        # Generate output path if not provided
        if output_path is None:
            output_path = tempfile.mktemp(suffix="_audit_logs.csv")

        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Headers
            writer.writerow([
                "ID", "Timestamp", "User ID", "Username", "Action", "Resource Type",
                "Resource ID", "Status", "IP Address", "Method", "Path",
                "Duration (ms)", "Error Message", "Checksum"
            ])

            # Data
            for log in logs:
                writer.writerow([
                    log.id,
                    log.timestamp.isoformat(),
                    log.user_id,
                    log.username,
                    log.action,
                    log.resource_type,
                    log.resource_id,
                    log.status,
                    log.ip_address,
                    log.request_method,
                    log.request_path,
                    log.duration_ms,
                    log.error_message,
                    log.checksum,
                ])

        return output_path

    def export_to_json(self, filters: AuditFilters, output_path: Optional[str] = None) -> str:
        """
        Export filtered audit logs to JSON.

        Args:
            filters: Search filters.
            output_path: Output file path. If None, creates temp file.

        Returns:
            Path to the generated JSON file.
        """
        # Get all matching logs
        export_filters = filters.model_copy()
        export_filters.page = 1
        export_filters.page_size = 100000

        logs, total = self.search_logs(export_filters)

        # Generate output path if not provided
        if output_path is None:
            output_path = tempfile.mktemp(suffix="_audit_logs.json")

        data = {
            "exported_at": datetime.now().isoformat(),
            "total_records": total,
            "filters": {
                "start_date": filters.start_date.isoformat() if filters.start_date else None,
                "end_date": filters.end_date.isoformat() if filters.end_date else None,
                "user_id": filters.user_id,
                "action": filters.action,
                "status": filters.status,
            },
            "logs": [log.model_dump(mode='json') for log in logs],
        }

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, default=str)

        return output_path


# Singleton instance
_audit_service: Optional[AuditService] = None


def get_audit_service(db_path: Optional[str] = None) -> AuditService:
    """Get or create the global AuditService instance."""
    global _audit_service
    if _audit_service is None:
        _audit_service = AuditService(db_path)
    return _audit_service
